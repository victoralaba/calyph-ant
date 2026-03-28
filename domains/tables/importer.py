# domains/tables/importer.py

"""
Data Import Engine.

Supports CSV, JSON, and Excel (.xlsx/.xls) imports into any PostgreSQL table.

Design principles:
  - Two-phase: preview first, commit second. Nothing touches the DB until
    the user explicitly confirms.
  - Type inference is conservative — when in doubt, use TEXT. Never lose data.
  - Column mapping is flexible — source columns can be renamed, skipped,
    or mapped to existing table columns.
  - Errors are collected per-row, not fatal. A bad row is reported and
    skipped; the rest still import.
  - Conflict handling: skip, overwrite (upsert), or error on PK collision.

Supported formats:
  - CSV (any delimiter, auto-detected; any encoding, auto-detected)
  - JSON (array of objects at root, or newline-delimited NDJSON)
  - Excel (.xlsx, .xls — first sheet by default, configurable)

Flow:
  1. parse_file()      → ParsedFile (rows, inferred_schema, warnings)
  2. preview_import()  → ImportPreview (sample rows, column map, type conflicts)
  3. execute_import()  → ImportResult (inserted, skipped, errors)
"""

from __future__ import annotations

import csv
import gzip
import io
import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from enum import Enum
from typing import Any
from uuid import UUID

import asyncpg
from loguru import logger


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

MAX_FILE_SIZE_BYTES = 50 * 1024 * 1024   # 50 MB hard limit
MAX_PREVIEW_ROWS = 50                     # Rows shown in preview
MAX_IMPORT_ROWS = 500_000                 # Hard cap per import
SAMPLE_ROWS_FOR_INFERENCE = 200           # Rows read for type inference

# Conflict strategies
class ConflictStrategy(str, Enum):
    skip      = "skip"       # Skip rows that conflict with existing PKs
    overwrite = "overwrite"  # UPDATE existing rows
    error     = "error"      # Abort entire import on first conflict


# ---------------------------------------------------------------------------
# Data models
# ---------------------------------------------------------------------------

@dataclass
class ColumnMapping:
    """Maps a source file column to a target table column."""
    source_name: str          # Column name in the file
    target_name: str          # Column name in the table (may differ)
    inferred_type: str        # Postgres type inferred from data
    skip: bool = False        # If True, this column is not imported
    sample_values: list[Any] = field(default_factory=list)
    null_count: int = 0
    non_null_count: int = 0


@dataclass
class ParsedFile:
    """Result of parsing a raw file into structured rows."""
    rows: list[dict[str, Any]]
    column_names: list[str]
    total_rows: int
    format: str                           # "csv" | "json" | "excel"
    encoding: str = "utf-8"
    delimiter: str = ","                  # CSV only
    sheet_name: str | None = None         # Excel only
    warnings: list[str] = field(default_factory=list)


@dataclass
class ImportPreview:
    """Preview returned to the user before they confirm import."""
    sample_rows: list[dict[str, Any]]     # First N rows after mapping
    column_mappings: list[ColumnMapping]
    total_rows: int
    format: str
    warnings: list[str]
    # Existing table columns for mapping UI
    table_columns: list[dict[str, Any]]
    # Whether the table exists yet (False = will be created)
    table_exists: bool


@dataclass
class ImportResult:
    """Final result after executing the import."""
    total: int
    inserted: int
    updated: int
    skipped: int
    errors: list[dict[str, Any]]          # [{row_index, error, data}]
    duration_ms: int
    table_created: bool = False


# ---------------------------------------------------------------------------
# Type inference
# ---------------------------------------------------------------------------

def _infer_pg_type(values: list[Any]) -> str:
    """
    Infer the most appropriate PostgreSQL column type from a sample of values.
    Conservative — prefers broader types over narrow ones to avoid data loss.
    """
    non_null = [v for v in values if v is not None and str(v).strip() not in ("", "null", "NULL", "None")]
    if not non_null:
        return "TEXT"

    # Try UUID
    if all(_looks_like_uuid(str(v)) for v in non_null):
        return "UUID"

    # Try boolean
    bool_true  = {"true", "yes", "1", "t", "y", "on"}
    bool_false = {"false", "no", "0", "f", "n", "off"}
    if all(str(v).lower() in (bool_true | bool_false) for v in non_null):
        return "BOOLEAN"

    # Try integer
    if all(_looks_like_int(v) for v in non_null):
        # Use BIGINT to be safe
        return "BIGINT"

    # Try numeric/decimal
    if all(_looks_like_numeric(v) for v in non_null):
        return "NUMERIC"

    # Try timestamp
    if all(_looks_like_timestamp(str(v)) for v in non_null):
        return "TIMESTAMPTZ"

    # Try date
    if all(_looks_like_date(str(v)) for v in non_null):
        return "DATE"

    # Try JSON
    json_count = sum(1 for v in non_null if _looks_like_json(str(v)))
    if json_count == len(non_null):
        return "JSONB"

    # Default
    max_len = max(len(str(v)) for v in non_null)
    if max_len <= 255:
        return "TEXT"
    return "TEXT"


def _looks_like_uuid(s: str) -> bool:
    return bool(re.match(
        r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$',
        s.strip(), re.IGNORECASE
    ))


def _looks_like_int(v: Any) -> bool:
    try:
        int(str(v).strip().replace(",", ""))
        return True
    except (ValueError, TypeError):
        return False


def _looks_like_numeric(v: Any) -> bool:
    try:
        Decimal(str(v).strip().replace(",", ""))
        return True
    except (InvalidOperation, TypeError):
        return False


_TIMESTAMP_PATTERNS = [
    r'^\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}',
    r'^\d{2}/\d{2}/\d{4} \d{2}:\d{2}',
    r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}',
]

def _looks_like_timestamp(s: str) -> bool:
    return any(re.match(p, s.strip()) for p in _TIMESTAMP_PATTERNS)


_DATE_PATTERNS = [
    r'^\d{4}-\d{2}-\d{2}$',
    r'^\d{2}/\d{2}/\d{4}$',
    r'^\d{2}-\d{2}-\d{4}$',
]

def _looks_like_date(s: str) -> bool:
    return any(re.match(p, s.strip()) for p in _DATE_PATTERNS)


def _looks_like_json(s: str) -> bool:
    s = s.strip()
    if not (s.startswith("{") or s.startswith("[")):
        return False
    try:
        json.loads(s)
        return True
    except (json.JSONDecodeError, ValueError):
        return False


# ---------------------------------------------------------------------------
# Value coercion
# ---------------------------------------------------------------------------

def _coerce_value(raw: Any, pg_type: str) -> Any:
    """
    Coerce a raw string/value to the appropriate Python type for asyncpg.
    Returns None for empty / null-like values.
    Raises ValueError if coercion fails (row will be reported as error).
    """
    if raw is None:
        return None
    s = str(raw).strip()
    if s in ("", "null", "NULL", "None", "NA", "N/A", "nan", "NaN"):
        return None

    t = pg_type.upper()

    if t in ("TEXT", "VARCHAR", "CHAR", "CITEXT"):
        return s

    if t in ("INTEGER", "INT", "INT4", "SMALLINT", "INT2"):
        return int(s.replace(",", ""))

    if t in ("BIGINT", "INT8"):
        return int(s.replace(",", ""))

    if t in ("NUMERIC", "DECIMAL", "REAL", "DOUBLE PRECISION", "FLOAT"):
        return float(s.replace(",", ""))

    if t == "BOOLEAN":
        if s.lower() in ("true", "yes", "1", "t", "y", "on"):
            return True
        if s.lower() in ("false", "no", "0", "f", "n", "off"):
            return False
        raise ValueError(f"Cannot coerce '{s}' to BOOLEAN")

    if t in ("UUID",):
        return s  # asyncpg handles UUID strings

    if t in ("DATE",):
        return s  # asyncpg parses date strings

    if t in ("TIMESTAMPTZ", "TIMESTAMP"):
        return s  # asyncpg parses timestamp strings

    if t in ("JSONB", "JSON"):
        if isinstance(raw, (dict, list)):
            return json.dumps(raw)
        try:
            json.loads(s)
            return s
        except json.JSONDecodeError:
            raise ValueError(f"Cannot coerce '{s[:50]}...' to JSON")

    # Fallback — return as string
    return s


# ---------------------------------------------------------------------------
# File parsers
# ---------------------------------------------------------------------------

def parse_csv(data: bytes) -> ParsedFile:
    """
    Parse CSV bytes into a ParsedFile.
    Auto-detects delimiter and encoding.
    """
    # Detect encoding
    try:
        import chardet
        detected = chardet.detect(data[:10000])
        encoding = detected.get("encoding") or "utf-8"
        if encoding.lower() in ("ascii", "windows-1252"):
            encoding = "utf-8"
    except ImportError:
        encoding = "utf-8"

    text = data.decode(encoding, errors="replace")
    sample = text[:4096]

    # Detect delimiter
    delimiter = ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        pass

    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    rows: list[dict[str, Any]] = []
    warnings: list[str] = []

    for i, row in enumerate(reader):
        if i >= MAX_IMPORT_ROWS:
            warnings.append(
                f"File truncated at {MAX_IMPORT_ROWS} rows. "
                "Only the first 500,000 rows will be imported."
            )
            break
        # Clean up keys
        cleaned = {k.strip(): (v.strip() if isinstance(v, str) else v) for k, v in row.items() if k}
        rows.append(cleaned)

    column_names = list(reader.fieldnames or [])

    return ParsedFile(
        rows=rows,
        column_names=column_names,
        total_rows=len(rows),
        format="csv",
        encoding=encoding,
        delimiter=delimiter,
        warnings=warnings,
    )


def parse_json(data: bytes) -> ParsedFile:
    """
    Parse JSON bytes. Supports:
    - Array of objects: [{...}, {...}]
    - Newline-delimited JSON (NDJSON): {...}\n{...}
    - Single object (wrapped in array automatically)
    """
    text = data.decode("utf-8", errors="replace").strip()
    warnings: list[str] = []
    rows: list[dict[str, Any]] = []

    # Try standard JSON first
    try:
        parsed = json.loads(text)
        if isinstance(parsed, list):
            rows = parsed
        elif isinstance(parsed, dict):
            # Check if it's a wrapper object with a data array
            for key in ("data", "rows", "records", "items", "results"):
                if key in parsed and isinstance(parsed[key], list):
                    rows = parsed[key]
                    warnings.append(f"Extracted rows from '{key}' key in JSON object.")
                    break
            else:
                rows = [parsed]
                warnings.append("Single JSON object wrapped in array for import.")
    except json.JSONDecodeError:
        # Try NDJSON
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError:
                warnings.append(f"Skipped invalid NDJSON line: {line[:50]}")

    if not rows:
        raise ValueError("No valid JSON rows found in file.")

    # Flatten rows — ensure all are dicts, stringify nested objects
    flat_rows: list[dict[str, Any]] = []
    all_keys: set[str] = set()
    for row in rows[:MAX_IMPORT_ROWS]:
        if not isinstance(row, dict):
            warnings.append(f"Skipped non-object row: {str(row)[:50]}")
            continue
        flat = {}
        for k, v in row.items():
            if isinstance(v, (dict, list)):
                flat[k] = json.dumps(v)
            else:
                flat[k] = v
            all_keys.add(k)
        flat_rows.append(flat)

    # Ensure all rows have all keys (fill missing with None)
    for row in flat_rows:
        for key in all_keys:
            row.setdefault(key, None)

    column_names = list(all_keys)
    if flat_rows:
        # Preserve insertion order from first row
        column_names = [k for k in flat_rows[0].keys()]

    return ParsedFile(
        rows=flat_rows,
        column_names=column_names,
        total_rows=len(flat_rows),
        format="json",
        warnings=warnings,
    )


def parse_excel(data: bytes, sheet_name: str | None = None) -> ParsedFile:
    """
    Parse Excel (.xlsx or .xls) bytes using openpyxl.
    Uses the first sheet by default.
    """
    try:
        import openpyxl
    except ImportError:
        raise ValueError(
            "openpyxl is required for Excel import. "
            "Install it with: pip install openpyxl"
        )

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    warnings: list[str] = []

    # Sheet selection
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        actual_sheet = sheet_name
    else:
        ws = wb.active
        actual_sheet: str = ws.title # type: ignore
        if sheet_name and sheet_name not in wb.sheetnames:
            warnings.append(
                f"Sheet '{sheet_name}' not found. "
                f"Using '{actual_sheet}'. "
                f"Available sheets: {', '.join(wb.sheetnames)}"
            )

    rows_iter = ws.iter_rows(values_only=True) # type: ignore
    header_row = next(rows_iter, None)
    if header_row is None:
        raise ValueError("Excel sheet is empty.")

    # Clean headers
    column_names = []
    for i, cell in enumerate(header_row):
        if cell is None:
            column_names.append(f"column_{i+1}")
        else:
            column_names.append(str(cell).strip())

    rows: list[dict[str, Any]] = []
    for i, row in enumerate(rows_iter):
        if i >= MAX_IMPORT_ROWS:
            warnings.append(f"File truncated at {MAX_IMPORT_ROWS} rows.")
            break
        # Skip entirely empty rows
        if all(v is None for v in row):
            continue
        row_dict: dict[str, Any] = {}
        for col_name, cell_val in zip(column_names, row):
            # Convert Excel date objects to strings
            if isinstance(cell_val, (datetime, date)):
                row_dict[col_name] = cell_val.isoformat()
            elif cell_val is not None:
                row_dict[col_name] = cell_val
            else:
                row_dict[col_name] = None
        rows.append(row_dict)

    wb.close()

    return ParsedFile(
        rows=rows,
        column_names=column_names,
        total_rows=len(rows),
        format="excel",
        sheet_name=actual_sheet,
        warnings=warnings,
    )


def parse_file(data: bytes, filename: str, sheet_name: str | None = None) -> ParsedFile:
    """
    Dispatch to the correct parser based on filename extension.
    Handles .gz compressed files transparently.
    """
    if len(data) > MAX_FILE_SIZE_BYTES:
        raise ValueError(
            f"File too large ({len(data) / 1024 / 1024:.1f} MB). "
            f"Maximum is {MAX_FILE_SIZE_BYTES // 1024 // 1024} MB."
        )

    name = filename.lower()

    # Decompress if gzipped
    if name.endswith(".gz"):
        data = gzip.decompress(data)
        name = name[:-3]

    if name.endswith(".csv") or name.endswith(".tsv") or name.endswith(".txt"):
        return parse_csv(data)
    elif name.endswith(".json") or name.endswith(".ndjson") or name.endswith(".jsonl"):
        return parse_json(data)
    elif name.endswith(".xlsx") or name.endswith(".xls"):
        return parse_excel(data, sheet_name=sheet_name)
    else:
        # Try to auto-detect
        text_sample = data[:512]
        try:
            decoded = text_sample.decode("utf-8", errors="strict")
            if decoded.strip().startswith(("[", "{")):
                return parse_json(data)
            else:
                return parse_csv(data)
        except UnicodeDecodeError:
            raise ValueError(
                f"Unsupported file format: '{filename}'. "
                "Supported formats: CSV, TSV, JSON, NDJSON, XLSX, XLS (optionally .gz compressed)"
            )


# ---------------------------------------------------------------------------
# Column mapping and type inference
# ---------------------------------------------------------------------------

def build_column_mappings(
    parsed: ParsedFile,
    table_columns: list[dict[str, Any]] | None = None,
) -> list[ColumnMapping]:
    """
    Build column mappings from parsed file columns to target table columns.

    If table_columns is provided (existing table), tries to match source
    columns to existing columns by name (case-insensitive).
    If no match found, the column will still be imported (table will need
    the column added, or user can skip it).
    """
    # Sample rows for inference
    sample = parsed.rows[:SAMPLE_ROWS_FOR_INFERENCE]

    # Build lookup for existing table columns
    existing_cols: dict[str, dict] = {}
    if table_columns:
        for col in table_columns:
            existing_cols[col["name"].lower()] = col

    mappings: list[ColumnMapping] = []
    for col_name in parsed.column_names:
        values = [row.get(col_name) for row in sample]
        non_null = [v for v in values if v is not None and str(v).strip() not in ("", "null", "NULL", "None")]
        null_count = len(values) - len(non_null)

        # Try to match to existing column
        existing = existing_cols.get(col_name.lower())
        if existing:
            pg_type = existing["data_type"]
        else:
            pg_type = _infer_pg_type(non_null)

        mapping = ColumnMapping(
            source_name=col_name,
            target_name=col_name,
            inferred_type=pg_type,
            sample_values=non_null[:5],
            null_count=null_count,
            non_null_count=len(non_null),
        )
        mappings.append(mapping)

    return mappings


# ---------------------------------------------------------------------------
# Preview
# ---------------------------------------------------------------------------

async def preview_import(
    pg_conn: asyncpg.Connection,
    parsed: ParsedFile,
    table_name: str,
    schema: str = "public",
    column_mappings: list[ColumnMapping] | None = None,
) -> ImportPreview:
    """
    Generate a preview of what the import will do.
    Does NOT modify the database.
    """
    warnings = list(parsed.warnings)

    # Check if table exists and get its columns
    table_exists = False
    table_columns: list[dict[str, Any]] = []
    try:
        rows = await pg_conn.fetch(
            """
            SELECT column_name AS name, data_type, is_nullable, column_default
            FROM information_schema.columns
            WHERE table_schema = $1 AND table_name = $2
            ORDER BY ordinal_position
            """,
            schema, table_name,
        )
        if rows:
            table_exists = True
            table_columns = [dict(r) for r in rows]
    except Exception as exc:
        logger.warning(f"Could not fetch table columns for preview: {exc}")

    # Build mappings if not provided
    if column_mappings is None:
        column_mappings = build_column_mappings(parsed, table_columns)

    # Generate sample rows using mappings
    active_mappings = [m for m in column_mappings if not m.skip]
    sample_rows: list[dict[str, Any]] = []
    for row in parsed.rows[:MAX_PREVIEW_ROWS]:
        mapped: dict[str, Any] = {}
        for mapping in active_mappings:
            raw = row.get(mapping.source_name)
            try:
                mapped[mapping.target_name] = _coerce_value(raw, mapping.inferred_type)
            except ValueError:
                mapped[mapping.target_name] = raw  # Show raw in preview
        sample_rows.append(mapped)

    # Warn about columns not in existing table
    if table_exists:
        existing_names = {c["name"].lower() for c in table_columns}
        for m in active_mappings:
            if m.target_name.lower() not in existing_names:
                warnings.append(
                    f"Column '{m.target_name}' does not exist in table "
                    f"'{table_name}'. It will be created automatically, "
                    "or you can skip it."
                )

    return ImportPreview(
        sample_rows=sample_rows,
        column_mappings=column_mappings,
        total_rows=parsed.total_rows,
        format=parsed.format,
        warnings=warnings,
        table_columns=table_columns,
        table_exists=table_exists,
    )


# ---------------------------------------------------------------------------
# Table creation from inferred schema
# ---------------------------------------------------------------------------

async def _ensure_table_exists(
    pg_conn: asyncpg.Connection,
    table_name: str,
    schema: str,
    column_mappings: list[ColumnMapping],
) -> bool:
    """
    Create the table if it doesn't exist.
    Returns True if table was created, False if it already existed.
    """
    active = [m for m in column_mappings if not m.skip]
    if not active:
        raise ValueError("No columns to create table with.")

    col_defs = []
    for m in active:
        safe_name = re.sub(r'[^\w]', '_', m.target_name)
        null_clause = ""  # Allow nulls by default on import
        col_defs.append(f'    "{safe_name}" {m.inferred_type}{null_clause}')

    cols_sql = ",\n".join(col_defs)
    create_sql = (
        f'CREATE TABLE IF NOT EXISTS "{schema}"."{table_name}" (\n'
        f"{cols_sql}\n"
        f");"
    )
    await pg_conn.execute(create_sql)
    return True


async def _add_missing_columns(
    pg_conn: asyncpg.Connection,
    table_name: str,
    schema: str,
    column_mappings: list[ColumnMapping],
) -> list[str]:
    """
    Add any columns from the mapping that don't exist in the table yet.
    Returns list of added column names.
    """
    existing_rows = await pg_conn.fetch(
        "SELECT column_name FROM information_schema.columns "
        "WHERE table_schema = $1 AND table_name = $2",
        schema, table_name,
    )
    existing = {r["column_name"].lower() for r in existing_rows}
    added: list[str] = []
    for m in column_mappings:
        if m.skip:
            continue
        if m.target_name.lower() not in existing:
            safe = re.sub(r'[^\w]', '_', m.target_name)
            await pg_conn.execute(
                f'ALTER TABLE "{schema}"."{table_name}" '
                f'ADD COLUMN IF NOT EXISTS "{safe}" {m.inferred_type}'
            )
            added.append(m.target_name)
    return added


# ---------------------------------------------------------------------------
# Core import execution
# ---------------------------------------------------------------------------

async def execute_import(
    pg_conn: asyncpg.Connection,
    parsed: ParsedFile,
    table_name: str,
    schema: str = "public",
    column_mappings: list[ColumnMapping] | None = None,
    conflict_strategy: ConflictStrategy = ConflictStrategy.skip,
    pk_column: str | None = None,
    create_table_if_missing: bool = True,
) -> ImportResult:
    """
    Execute the actual import into PostgreSQL.

    Uses COPY-style bulk insertion via asyncpg executemany for performance,
    with per-row fallback when individual rows fail.

    Args:
        pg_conn:                 Live asyncpg connection to target DB
        parsed:                  ParsedFile from parse_file()
        table_name:              Target table name
        schema:                  Target schema (default: public)
        column_mappings:         Column mapping config (built automatically if None)
        conflict_strategy:       How to handle PK conflicts
        pk_column:               Primary key column name (for upsert)
        create_table_if_missing: Auto-create table if it doesn't exist
    """
    import time
    start = time.monotonic()

    # Build mappings if not provided
    if column_mappings is None:
        table_col_rows = await pg_conn.fetch(
            "SELECT column_name AS name, data_type FROM information_schema.columns "
            "WHERE table_schema = $1 AND table_name = $2",
            schema, table_name,
        )
        table_columns = [dict(r) for r in table_col_rows]
        column_mappings = build_column_mappings(parsed, table_columns)

    active_mappings = [m for m in column_mappings if not m.skip]
    if not active_mappings:
        raise ValueError("No columns selected for import.")

    table_created = False

    # Check table existence
    exists_row = await pg_conn.fetchval(
        "SELECT 1 FROM information_schema.tables WHERE table_schema = $1 AND table_name = $2",
        schema, table_name,
    )
    table_exists = exists_row is not None

    if not table_exists:
        if create_table_if_missing:
            await _ensure_table_exists(pg_conn, table_name, schema, active_mappings)
            table_created = True
        else:
            raise ValueError(
                f"Table '{schema}.{table_name}' does not exist. "
                "Set create_table_if_missing=True to auto-create it."
            )
    else:
        # Add any new columns
        await _add_missing_columns(pg_conn, table_name, schema, active_mappings)

    # Prepare column names and SQL
    col_names = [f'"{m.target_name}"' for m in active_mappings]
    cols_sql = ", ".join(col_names)
    placeholders = ", ".join(f"${i+1}" for i in range(len(active_mappings)))

    if conflict_strategy == ConflictStrategy.skip and pk_column:
        insert_sql = (
            f'INSERT INTO "{schema}"."{table_name}" ({cols_sql}) '
            f"VALUES ({placeholders}) "
            f'ON CONFLICT ("{pk_column}") DO NOTHING'
        )
    elif conflict_strategy == ConflictStrategy.overwrite and pk_column:
        set_parts = [
            f'{col_names[i]} = EXCLUDED.{col_names[i]}'
            for i in range(len(active_mappings))
            if active_mappings[i].target_name != pk_column
        ]
        insert_sql = (
            f'INSERT INTO "{schema}"."{table_name}" ({cols_sql}) '
            f"VALUES ({placeholders}) "
            f'ON CONFLICT ("{pk_column}") DO UPDATE SET '
            + ", ".join(set_parts)
        )
    else:
        insert_sql = (
            f'INSERT INTO "{schema}"."{table_name}" ({cols_sql}) '
            f"VALUES ({placeholders})"
        )

    # Execute — try bulk first, fall back to per-row
    inserted = 0
    updated = 0
    skipped = 0
    errors: list[dict[str, Any]] = []

    # Build value tuples
    value_tuples: list[tuple] = []
    coerce_errors: list[dict] = []

    for row_idx, row in enumerate(parsed.rows):
        try:
            values = tuple(
                _coerce_value(row.get(m.source_name), m.inferred_type)
                for m in active_mappings
            )
            value_tuples.append(values)
        except (ValueError, Exception) as exc:
            coerce_errors.append({
                "row_index": row_idx + 1,
                "error": f"Type conversion failed: {exc}",
                "data": {m.source_name: row.get(m.source_name) for m in active_mappings},
            })

    errors.extend(coerce_errors)

    # Try bulk executemany
    try:
        async with pg_conn.transaction():
            await pg_conn.executemany(insert_sql, value_tuples)
        inserted = len(value_tuples)

    except Exception as bulk_exc:
        # Bulk failed — fall back to row-by-row
        logger.info(
            f"Bulk insert failed ({bulk_exc}), falling back to row-by-row import"
        )

        for row_idx, values in enumerate(value_tuples):
            try:
                result = await pg_conn.execute(insert_sql, *values)
                # asyncpg returns e.g. "INSERT 0 1" or "UPDATE 1"
                if "UPDATE" in result:
                    updated += 1
                elif "INSERT 0 0" in result:
                    skipped += 1
                else:
                    inserted += 1
            except asyncpg.UniqueViolationError:
                if conflict_strategy == ConflictStrategy.error:
                    raise ValueError(
                        f"PK conflict on row {row_idx + 1}. "
                        "Use conflict_strategy='skip' or 'overwrite' to handle conflicts."
                    )
                skipped += 1
            except Exception as row_exc:
                errors.append({
                    "row_index": row_idx + 1,
                    "error": str(row_exc),
                    "data": {
                        m.source_name: values[i]
                        for i, m in enumerate(active_mappings)
                    },
                })

    duration_ms = int((time.monotonic() - start) * 1000)

    return ImportResult(
        total=len(parsed.rows),
        inserted=inserted,
        updated=updated,
        skipped=skipped,
        errors=errors,
        duration_ms=duration_ms,
        table_created=table_created,
    )


# ---------------------------------------------------------------------------
# Export helpers (for round-trip testing and download)
# ---------------------------------------------------------------------------

def rows_to_csv(rows: list[dict[str, Any]], column_names: list[str]) -> bytes:
    """Convert rows to CSV bytes for download."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=column_names, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return buf.getvalue().encode("utf-8")


def rows_to_json(rows: list[dict[str, Any]]) -> bytes:
    """Convert rows to JSON bytes for download."""
    return json.dumps(rows, default=str, indent=2).encode("utf-8")
